import copy
import logging
import math

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string

from deca.ff_adf import Adf, AdfValue
from deca.file import ArchiveFile
from modbuilder import mods
from modbuilder.logging_config import get_logger

logger = get_logger(__name__)


def deserialize_adf(filename: str, modded: bool = True) -> Adf:
  file = mods.get_modded_file(filename) if modded else mods.get_org_file(filename)
  adf = Adf()
  with ArchiveFile(open(file, 'rb')) as f:
    adf.deserialize(f)
  return adf


class XlsxCell:
  __slots__ = (
    'coordinates',              # coordinates of cell in sheet
    'sheet_name',               # sheet that cell is located on
    'sheet_index',              # index of the sheet in the extracted ADF file
    'src_filename',             # file containing the sheet data
    'index',                    # index of the cell from all cells in the sheet
    'definition_index',         # index in the list of cell definitions
    'definition_index_offset',  # offset where "defintiion_index" is stored
    'value_index',              # index of the value in the data array
    'value_index_offset',       # offset where "value_index" is stored
    'value',                    # value of the cell in the parsed ADF
    'value_offset',             # offset where the value is stored in the data array
    'data_type',                # type of data (0=bool, 1=str, 2=float)
    'data_array_name',          # name of the array the value is stored in
    'attribute_index',          # attribute_index from cell definition
    'desired_value',            # ideal value after transform is applied
    'desired_data_type',        # data type of the ideal value
    'desired_data_array_name',  # name of the array the desired value is stored in
  )

  coordinates: str
  sheet_name: str
  sheet_index: int
  src_filename: str
  index: int
  definition_index: int
  definition_index_offset: int
  value_index: int
  value_index_offset: int
  value: any
  value_offset: int
  data_type: int
  data_array_name: str
  attribute_index: int
  desired_value: any
  desired_data_type: int
  desired_data_array_name: str

  def __init__(self, src_filename: str, extracted_adf: Adf, update_data: dict) -> 'XlsxCell':
    self.src_filename = src_filename
    self.coordinates = update_data["coordinates"]
    self.sheet_name = update_data["sheet"]
    adf_values = extracted_adf.table_instance_full_values[0].value
    sheet, self.sheet_index = get_sheet(extracted_adf, self.sheet_name)
    if sheet is None:
      raise ValueError(f'Unable to find sheet "{self.sheet_name}" in file "{src_filename}"')
    self._get_value_and_offsets(adf_values, sheet, src_filename)
    if "value" in update_data:  # update_data without "value" can be used to extract a cell from an ADF
      self._format_desired_data(update_data)

  def _get_value_and_offsets(self, adf_values: dict[str, AdfValue], sheet: AdfValue, src_filename: str) -> None:
    col_str, row = coordinate_from_string(self.coordinates)
    col = column_index_from_string(col_str)
    self.index = ( ( row - 1 ) * sheet["Cols"].value ) + col - 1  # -1 because spreadsheets start at A1 but lists start at 0

    self.definition_index = int(sheet["CellIndex"].value[self.index])
    self.definition_index_offset = int(sheet["CellIndex"].data_offset + ( self.index * 4 ))

    cell_definition = adf_values["Cell"].value[self.definition_index].value
    self.value_index = int(cell_definition["DataIndex"].value)
    self.value_index_offset = int(cell_definition["DataIndex"].data_offset)
    self.data_type = int(cell_definition["Type"].value)
    self.attribute_index = int(cell_definition["AttributeIndex"].value)

    self.value = None
    if self.data_type == 0:
      self.data_array_name = "BoolData"
      self.value = adf_values[self.data_array_name].value[self.value_index]
      self.value_offset = adf_values[self.data_array_name].data_offset + ( self.value_index * 4 )
    if self.data_type == 1:
      self.data_array_name = "StringData"
      self.value = adf_values[self.data_array_name].value[self.value_index].value
      self.value_offset = adf_values[self.data_array_name].value[self.value_index].data_offset
    if self.data_type == 2:
      self.data_array_name = "ValueData"
      self.value = adf_values[self.data_array_name].value[self.value_index]
      self.value_offset = adf_values[self.data_array_name].data_offset + ( self.value_index * 4 )
    if self.value is None:
      raise ValueError(f'Unable to find cell "{self.coordinates}" on sheet "{self.sheet_name}" in file "{src_filename}"')

  def _format_desired_data(self, update_data: dict) -> dict:
    transform = update_data.get("transform")
    if transform == "multiply":
      self.desired_value = self.value * update_data["value"]
    elif transform == "add":
      self.desired_value = self.value + update_data["value"]
    else:
      self.desired_value = update_data["value"]

    if isinstance(self.desired_value, bool):
      self.desired_data_type = 0
      self.desired_data_array_name = "BoolData"
    if isinstance(self.desired_value, str):
      self.desired_value = self.desired_value.encode("utf-8")  # for comparison against encoded values in StringData
      self.desired_data_type = 1
      self.desired_data_array_name = "StringData"
    if isinstance(self.desired_value, float) or isinstance(self.desired_value, int):
      self.desired_value = float(self.desired_value)
      self.desired_data_type = 2
      self.desired_data_array_name = "ValueData"


def get_sheet(extracted_adf: Adf, sheet_name: str) -> tuple[AdfValue, int]:
  adf_values = extracted_adf.table_instance_full_values[0].value
  for i, sheet in enumerate(adf_values["Sheet"].value):
    if sheet.value["Name"].value.decode("utf-8") == sheet_name:
      return sheet.value, i
  return None, None


def process_cell_update(cell: XlsxCell, extracted_adf: Adf, skip_add_data: bool = False, allow_new_data: bool = False, force: bool = False) -> list[dict]:
  # skip_add_data = completely skip attempts to data to the file - just re-use existing values. This can be slow on large spreadsheets (animal_senses.bin)
  # allow_new_data = enablie adding new String/Value data and cell definitions. Some files will cause crashes with added data
  adf_values = extracted_adf.table_instance_full_values[0].value
  logger.debug("")
  logger.debug(f'Cell = Coordinates: {cell.coordinates}   Sheet: {cell.sheet_name}   Value: {cell.value}   Data Type: {cell.data_type}   Value Index: {cell.value_index}')
  logger.debug(f'Definition = Index: {cell.definition_index}   AttributeIndex: {cell.attribute_index}')
  logger.debug(f'Desired = Value: {cell.desired_value}   Data Type: {cell.desired_data_type}')

  # 0. If current cell already has desired value and is the correct type, don't change anything
  if cell.value == cell.desired_value and cell.data_type == cell.desired_data_type:
    logger.debug('0. Current value and data type match. No changes to apply')
    return []

  # Try to find a safe way to overwrite data or repoint cell definitions to preferred values
  # 1. Check if the desired value is already in the data array. If it is then work with it
  if is_desired_value_in_data_array(adf_values, cell):
    logger.debug(f'1. Desired value {cell.desired_value} is in data array {cell.desired_data_array_name}')
    if (file_updates := use_value_from_data_array(adf_values, cell)):
      return file_updates

  # 2. Desired value does not exist in the data array. See if we can write it in without conflicts
  if not skip_add_data:
    logger.debug(f'2. Desired value {cell.desired_value} is NOT in the data array {cell.desired_data_array_name}')
    if (file_updates := write_value_to_data_array(extracted_adf, cell)):
      return file_updates
    if allow_new_data:
      if (file_updates := add_new_value_to_data_array(extracted_adf, cell)):
        return file_updates

  # 3. Cannot update the value of the specified cell to match the exact desired value without affecting other cells
  logger.debug(f'3. Cannot update cell {cell.coordinates} to exact value {cell.desired_value} without affecting other cells')
  # If we were told to force the exact value then overwrite the value in the array
  if force:
    logger.debug(f'FORCE: Overwriting {cell.data_array_name} value at index {cell.value_index} with {cell.desired_value}')
    return overwrite_value(extracted_adf, cell)

  # Locate the closest value in the data array and see if we can work with that. This only works for ValueData cells
  if cell.data_array_name == "ValueData":
    if (file_updates := use_closest_value_in_array(adf_values, cell)):
      return file_updates

  # If you got all the way down here then you're out of luck for now.
  raise NotImplementedError(f'Unable to update cell {cell.coordinates} with the desired value {cell.desired_value}. Exiting...')


def is_desired_value_in_data_array(adf_values: dict[str, AdfValue], cell: XlsxCell) -> bool:
  if cell.data_array_name in ["BoolData", "ValueData"]:  # values stored directly in array
    return cell.desired_value in adf_values[cell.desired_data_array_name].value
  if cell.data_array_name == "StringData":  # StringData is an array of AdfValues
    array_values = [s.value for s in adf_values["StringData"].value]
    return cell.desired_value in array_values
  return False


def use_value_from_data_array(adf_values: dict[str, AdfValue], cell: XlsxCell) -> list[dict]:
  # Value may exist more than once in array
  if cell.desired_data_array_name == "StringData":
    desired_value_indexes = [i for i, data in enumerate(adf_values[cell.desired_data_array_name].value) if data.value == cell.desired_value]
  else:
    desired_value_indexes = [i for i, value in enumerate(adf_values[cell.desired_data_array_name].value) if value == cell.desired_value]

  file_updates = []
  # 1a. Point current cell at a different definition that references the desired value
  #     The definition must also match our cell's attributes (text and background color)
  cell_defs_with_matching_value = find_cell_definitions(adf_values, cell.desired_data_type, desired_value_indexes)
  matching_definition = None
  if cell_defs_with_matching_value:
    logger.debug(f"   Found {len(cell_defs_with_matching_value)} cells with the desired value")
    cell_defs_with_matching_attribute = [def_tuple for def_tuple in cell_defs_with_matching_value if def_tuple[1]["AttributeIndex"].value == cell.attribute_index]
    logger.debug(f"   Found {len(cell_defs_with_matching_attribute)} cells with matching attributes")
    if cell_defs_with_matching_attribute:
      matching_definition = cell_defs_with_matching_attribute.pop()
    if matching_definition:
      file_updates.append({"offset": cell.definition_index_offset, "value": matching_definition[0]})
      adf_values["Sheet"].value[cell.sheet_index].value["CellIndex"].value[cell.index] = matching_definition[0]
      logger.debug(f'1a. Cell definition {matching_definition[0]} points at the desired value. Updating cell to use that definition.')
      return file_updates

  # 1b. Check if any other cells use the same definition as our cell
  #     If not, point current cell definition at desired value
  cells_with_same_definition = find_cells(adf_values, definition_indexes=[cell.definition_index], ignore_cell=cell)
  logger.debug(f'1b. {len(cells_with_same_definition)} other cells point at the same definition.')
  if not cells_with_same_definition:
    file_updates.append({"offset": cell.value_index_offset, "value": desired_value_indexes[0]})
    adf_values["Cell"].value[cell.definition_index].value["DataIndex"].value = desired_value_indexes[0]
    logger.debug(f'1b. No other cells share the same definition. Repointing defintiion at desired value {cell.desired_value} at index {desired_value_indexes[0]}.')
    return file_updates

  # 1c. Overwrite an unused cell definition (if one exists) to point it at the desired value in the data array
  #     Then point our cell at the customized cell definition
  unused_definition_indexes = get_unused_cell_def_indexes(adf_values)
  if unused_definition_indexes:
    unused_definition_index = unused_definition_indexes.pop()
    logger.debug(f'1c. Overwriting unused cell definition {unused_definition_index} to point at desired value {cell.desired_value} at index {desired_value_indexes[0]}')
    unused_definition = adf_values["Cell"].value[unused_definition_index].value
    # Point unused cell definition at the desired value index
    file_updates.append({"offset": unused_definition["DataIndex"].data_offset, "value": desired_value_indexes[0]})
    adf_values["Cell"].value[unused_definition_index].value["DataIndex"].value = desired_value_indexes[0]
    # Point our cell at the updated definition
    file_updates.append({"offset": cell.definition_index_offset, "value": unused_definition_index})
    adf_values["Sheet"].value[cell.sheet_index].value["CellIndex"].value[cell.index] = unused_definition_index
    return file_updates
  # 1z. Unable to use the existing value in the array. Find another way to do it
  return None


def write_value_to_data_array(extracted_adf: Adf, cell: XlsxCell) -> list[dict]:
  adf_values = extracted_adf.table_instance_full_values[0].value
  file_updates = []
  logger.debug(f"2a. Attempting to writing value {cell.desired_value} to the data array")
  # 2a. Check if we can overwrite our current value
  #     Check if any other cells in the entire sheet point at the same value as our cell
  #     There might be a cell index that points at our cell but it could be unused
  cells_with_shared_value = find_cells(adf_values, data_type=cell.data_type, value_index=cell.value_index, ignore_cell=cell)
  logger.debug(f"  - {len(cells_with_shared_value)} other cells point at the same value.")
  if not cells_with_shared_value:  # if none are found then overwrite our current value in the data array
    logger.debug(f'  - Overwriting {cell.desired_data_array_name} value at index {cell.value_index} to new value {cell.desired_value}')
    file_updates.extend(overwrite_value(extracted_adf, cell))
    return file_updates
  else:
    logger.debug(f'  - Cannot overwrite data at index {cell.value_index} directly without affecting other cells.')

  # 2b/c. Try to find an unused item in the data array to overwrite
  unused_values = get_unused_values(adf_values, cell.desired_data_array_name, cell.desired_data_type)
  # 2b. If one exists, check if any other cells share our definition
  #     If not, overwrite the value and point our definition at the new value
  if unused_values:
    logger.debug(f'2b. Found {len(unused_values)} unused values in {cell.desired_data_array_name}')
    if logger.isEnabledFor(logging.DEBUG):  # skip this loop if logger is not in DEBUG mode
      for unused_value in unused_values:
        logger.debug(f'  - Index: {unused_value["index"]}   Value: {unused_value["value"]}')
    if not find_cells(adf_values, definition_indexes=[cell.definition_index], ignore_cell=cell):
      # Overwrite the unused data array item
      unused_value = unused_values.pop()
      logger.debug(f'  - Overwriting unused {cell.desired_data_array_name} value at index {unused_value["index"]} to new value {cell.desired_value}')
      file_updates.extend(overwrite_value(extracted_adf, cell, unused_value))
      file_updates.append({"offset": cell.value_index_offset, "value": unused_value["index"]})
      adf_values["Cell"].value[cell.definition_index].value["DataIndex"].value = unused_value["index"]
      logger.debug(f'  - Pointing cell {cell.coordinates} definition {cell.definition_index} at value index {unused_value["index"]} with value {cell.desired_value}')
      return file_updates

  # 2c. There's still an unused data array item we can overwrite
  #     Check if there is an unused cell definition that we can repurpose
  #     If so. overwrite the unused data item, repoint the unused definition, and repoint our cell at that definition
  unused_definition_indexes = get_unused_cell_def_indexes(adf_values)
  if unused_values and unused_definition_indexes:
    unused_value = unused_values.pop()
    unused_definition_index = unused_definition_indexes.pop()
    unused_definition = adf_values["Cell"].value[unused_definition_index].value

    # Overwrite the unused data array item
    logger.debug(f'2c. Overwriting unused {cell.desired_data_array_name} value at index {unused_value["index"]} to new value {cell.desired_value}')
    file_updates.extend(overwrite_value(extracted_adf, cell, unused_value))

    # Point the unused cell definition at the new data array item
    logger.debug(f'  - Overwriting unused cell definition {unused_definition_index} to point at value index {unused_value["index"]}')
    file_updates.append({"offset": unused_definition["DataIndex"].data_offset, "value": unused_value["index"]})
    adf_values["Cell"].value[unused_definition_index].value["DataIndex"].value = unused_value["index"]
    # Copy our cell definition type and attribute to the unused cell definition
    file_updates.append({"offset": unused_definition["Type"].data_offset, "value": cell.data_type})
    file_updates.append({"offset": unused_definition["AttributeIndex"].data_offset, "value": cell.attribute_index})
    adf_values["Cell"].value[unused_definition_index].value["Type"].value = cell.data_type
    adf_values["Cell"].value[unused_definition_index].value["AttributeIndex"].value = cell.attribute_index
    # Point our cell at the definition
    logger.debug(f'  - Pointing cell {cell.coordinates} at definition {unused_definition_index} with value {cell.desired_value}')
    file_updates.append({"offset": cell.definition_index_offset, "value": unused_definition_index})
    adf_values["Sheet"].value[cell.sheet_index].value["CellIndex"].value[cell.index] = unused_definition_index
    return file_updates

  # 2z. Unable to update a value in the array. Find another way to do it
  return None


def overwrite_value(extracted_adf: Adf, cell: XlsxCell, target: dict = None) -> list[tuple[int, any]]:
  file_updates = []
  adf_values = extracted_adf.table_instance_full_values[0].value
  if not target:
    target = {"index": cell.value_index, "value": cell.value, "offset": cell.value_offset}

  if cell.desired_data_array_name == "ValueData":
    # ValueData are 4-byte float and can be directly overwritten
    file_updates.append({"offset": target["offset"], "value": cell.desired_value})
    adf_values["ValueData"].value[target["index"]] = cell.desired_value
    return file_updates
  if cell.desired_data_array_name == "StringData":
    stringdata = adf_values["StringData"]
    # StringData values are arbitrary lengths. Their data_offsets are stored in info_offset headers before the array
    # Strings are byte-aligned and have data_offsets divisible by 8 with at least 1 byte of padding between strings
    if target["index"] == len(stringdata.value):  # last value in array = padded until start of ValueData
      padding = adf_values["ValueData"].data_offset - (target["offset"] + len(target["value"]))
    else:  # padded until start of next string
      padding = stringdata.value[target["index"] + 1].data_offset - (target["offset"] + len(target["value"]))
    old_str_bytes = target["value"] + b'\x00' * padding
    new_str_bytes = bytes(mods.create_bytearray(cell.desired_value, "string"))
    difference = len(new_str_bytes) - len(old_str_bytes)
    logger.debug(f"OLD_STR: {old_str_bytes}   NEW_STR: {new_str_bytes}   DIFF: {difference}")
    if difference <= 0:  # can overwrite in place. add extra padding to ensure old value is totally overwritten
      new_str_bytes += b'\x00' * difference
      file_updates.append({"offset": target["offset"], "value": new_str_bytes})
      stringdata.value[target["index"]].value = cell.desired_value
      logger.debug(f"   Overwriting value '{target["value"]}' with new value {cell.desired_value} at offset {target["offset"]}")
      return file_updates
    else:  # new string is too long. Insert in the new string and shift all data the comes later in the file
      # update headers and instance offsets
      file_updates.extend(mods.update_non_instance_offsets(extracted_adf, difference))
      file_updates.extend(update_instance_offsets(extracted_adf, difference, "StringData"))
      # increase offsets of all string data values later in the array
      file_updates.extend(update_stringdata_offsets(extracted_adf, difference, index=target["index"]))
      logger.debug(f"   Overwriting value '{target["value"]}' with new value {cell.desired_value} at offset {target["offset"]}")
      # add to array and increase array length in header
      file_updates.append({
        "offset": target["offset"],
        "value": new_str_bytes,
        "transform": "insert",
        "bytes_to_remove": len(old_str_bytes),
      })
      stringdata.value[target["index"]].value = cell.desired_value
      return file_updates


def add_new_value_to_data_array(extracted_adf: Adf, cell: XlsxCell) -> list[dict]:
  adf_values = extracted_adf.table_instance_full_values[0].value
  file_updates = []
  logger.debug(f'2d. Adding value {cell.desired_value} to the data array')
  if cell.desired_data_array_name == "ValueData":
    file_updates.extend(add_float_to_valuedata(extracted_adf, cell.desired_value))
    new_value_index = len(adf_values["ValueData"].value) - 1
  elif cell.desired_data_array_name == "StringData":
    file_updates.extend(add_string_to_stringdata(extracted_adf, cell.desired_value))
    new_value_index = len(adf_values["StringData"].value) - 1
  else:
    return []
  # Check if any other cells share our definition
  if not find_cells(adf_values, definition_indexes=[cell.definition_index], ignore_cell=cell):
    # Point our cell definition at new array value
    logger.debug(f'  - Pointing cell {cell.coordinates} definition {cell.definition_index} at value index {new_value_index} with value {cell.desired_value}')
    file_updates.append({"offset": cell.value_index_offset, "value": new_value_index})
    adf_values["Cell"].value[cell.definition_index].value["DataIndex"].value = new_value_index
  else:  # Check if there is an unused definition we can overwrite
    if unused_definition_indexes:= get_unused_cell_def_indexes(adf_values):
      logger.debug(f"  - Found {len(unused_definition_indexes)} unused definitions: {unused_definition_indexes}")
      def_index = unused_definition_indexes.pop()
      # Update the Type of the unused cell definition
      file_updates.append({"offset": adf_values["Cell"].value[def_index].value["Type"].data_offset, "value": cell.desired_data_type})
      adf_values["Cell"].value[def_index].value["Type"].value = cell.desired_data_type
      # Update the AttributeIndex of the unused cell definition
      file_updates.append({"offset": adf_values["Cell"].value[def_index].value["AttributeIndex"].data_offset, "value": cell.attribute_index})
      adf_values["Cell"].value[def_index].value["AttributeIndex"].value = cell.attribute_index
      # Point unused cell definition at the desired value index
      file_updates.append({"offset": adf_values["Cell"].value[def_index].value["DataIndex"].data_offset, "value": new_value_index})
      adf_values["Cell"].value[def_index].value["DataIndex"].value = new_value_index
      # Point our cell at the updated definition
      file_updates.append({"offset": cell.definition_index_offset, "value": def_index})
      adf_values["Sheet"].value[cell.sheet_index].value["CellIndex"].value[cell.index] = def_index
      logger.debug(f'  - Overwriting unused cell definition {def_index} to point at desired value {cell.desired_value} at index {new_value_index}')
    else:  # Create a new one with our desired specs
      def_index = len(adf_values["Cell"].value)
      logger.debug(f'  - Creating new cell definition at index {def_index} (Type: {cell.desired_data_type}  DataIndex: {new_value_index}  AttributeIndex: {cell.attribute_index})')
      file_updates.extend(add_cell_definition(extracted_adf, cell, new_value_index))
    # Point our cell at the selected definition
    logger.debug(f'  - Pointing cell {cell.coordinates} at definition {def_index} with value {cell.desired_value}')
    file_updates.append({"offset": cell.definition_index_offset, "value": def_index})
    adf_values["Sheet"].value[cell.sheet_index].value["CellIndex"].value[cell.index] = def_index
    file_updates.append({"offset": 4, "value": 3})  # ADFv3 to prevent crash on load
  return file_updates


def use_closest_value_in_array(adf_values: dict[str, AdfValue], cell: XlsxCell) -> list[tuple[int, int]]:
  file_updates = []
  # Find the closest value in the array
  closest_value_index, closest_value = find_closest_value(adf_values["ValueData"].value, cell.desired_value)
  logger.debug(f'  Closest value in data array is {closest_value} at index {closest_value_index}')
  # 3a. Check if any other cells are using the same definition as our cell.
  #     If not, point our cell definition at the closest value in the array
  cells_with_same_definition = find_cells(adf_values, definition_indexes=[cell.definition_index], ignore_cell=cell)
  logger.debug(f'3a. {len(cells_with_same_definition)} cells point at the same definition as our cell.')
  if not cells_with_same_definition:
    file_updates.append({"offset": cell.value_index_offset, "value": closest_value_index})
    adf_values["Cell"].value[cell.definition_index].value["DataIndex"].value = closest_value_index
    logger.debug(f'  - Pointing cell {cell.coordinates} definition {cell.definition_index} with closest value {closest_value}')
    return file_updates
  # 3b. Check if any other definitions are pointing at the closest value
  #     If one is found then point our cell at that definition
  defs_with_closest_value = find_cell_definitions(adf_values, cell.data_type, [closest_value_index])
  logger.debug(f'  - {len(defs_with_closest_value)} cell definitions point at a definition with the closest value {closest_value}: {[d[0] for d in defs_with_closest_value]}')
  if defs_with_closest_value:
    defs_with_same_attributes = [(i,d) for (i,d) in defs_with_closest_value if d["AttributeIndex"].value == cell.attribute_index]
    if defs_with_same_attributes:
      logger.debug(f'  - {len(defs_with_same_attributes)} cell definitions have the same attribute ({cell.attribute_index}): {[d[0] for d in defs_with_same_attributes]}.')
      chosen_definition = defs_with_same_attributes.pop()
    else:
      chosen_definition = defs_with_closest_value.pop()
    file_updates.append({"offset": cell.definition_index_offset, "value": chosen_definition[0]})
    adf_values["Sheet"].value[cell.sheet_index].value["CellIndex"].value[cell.index] = chosen_definition[0]
    logger.debug(f'3b. Pointing cell {cell.coordinates} at definition {chosen_definition[0]} with closest value {closest_value}')
    return file_updates
  # 3c. Point our cell definition at the closest value
  #     This will have unintended consequences since other cells use this cell definition
  file_updates.append({"offset": cell.value_index_offset, "value": closest_value_index})
  adf_values["Cell"].value[cell.definition_index].value["DataIndex"].value = closest_value_index
  logger.debug(f'3c. Pointing cell {cell.coordinates} definition {cell.definition_index} at value index {closest_value_index} with closest value {closest_value}')
  return file_updates


def update_instance_offsets(extracted_adf: Adf, added_size: int, data_array_name: str) -> list[dict]:
  file_updates = []
  adf_values = extracted_adf.table_instance_full_values[0].value
  # only update arrays that are later in the file. Thankfully dictionaries are ordered
  array_names = list(adf_values.keys())
  if data_array_name in array_names:
    i = array_names.index(data_array_name)
    arrays_to_update = array_names[i + 1:]
  for array_name in arrays_to_update:
    file_updates.extend(update_array_offsets(extracted_adf, added_size, array_name))
  return file_updates


def update_array_offsets(extracted_adf: Adf, added_size: int, array_name: str) -> list[dict]:
  file_updates = []
  instance_offset = extracted_adf.table_instance[0].offset
  adf_values = extracted_adf.table_instance_full_values[0].value
  if adf_values[array_name].data_offset != instance_offset:  # an empty array has the same offset as the instance
    logger.debug(f"  Updating {array_name} array offsets by {added_size}")
    file_updates.append({
      "offset": adf_values[array_name].info_offset,
      "value": added_size,
      "transform": "add"
    })
    adf_values[array_name].data_offset += added_size
    # StringData stores strings of arbitrary length in AdfValue objects that have separate info offsets
    if array_name == "StringData":
      file_updates.extend(update_stringdata_offsets(extracted_adf, added_size))
    # Attribute array contains AdfValues with nested dictionaries of AdfValues
    # These just point to offsets in other arrays. Update the extracted Adf but don't write anything back to the file
    if array_name == "Attribute":
      for i in adf_values[array_name].value:
        i.info_offset += added_size
        i.data_offset += added_size
        for k,v in i.value.items():
          v.info_offset += added_size
          v.data_offset += added_size
    if array_name == "DateData":  # never seen DateData and don't know what it looks like. If we find data then error out
      if adf_values[array_name].value:
        raise NotImplementedError("Updating DateData is not supported")

  else:
    logger.debug(f"  No offset update for array {array_name}")
  return file_updates


def update_stringdata_offsets(extracted_adf: Adf, added_size: int, index: int = -1) -> list[dict]:
  # StringData values are arbitrary lengths. Their data_offsets are stored at the info_offset in StringData headers before the string array
  file_updates = []
  stringdata_array = extracted_adf.table_instance_full_values[0].value["StringData"].value
  if logger.isEnabledFor(logging.DEBUG):  # skip this if logger is in DEBUG mode
    i_to_update = [i for i in range(len(stringdata_array)) if i > index]
    info_text = "info and " if index >= 0 else ""
    logger.debug(f"  Updating {len(i_to_update)} StringData {info_text}data offsets by {added_size}")
    # logger.debug(i_to_update)
  for i, s in enumerate(stringdata_array):
    if (
      index == -1  # data was added before StringData = shift all offsets
      or index > len(stringdata_array)  # new string was added and the end of array = adjust for new header uint32
      or i > index  # existing string was overwritten = shift offsets for all strings later in the array
    ):
        file_updates.append({"offset": s.info_offset, "value": added_size, "transform": "add"})
        s.data_offset += added_size
        if index == -1:  # only need to shift info offsets if data was added before the StringData array
            s.info_offset += added_size
  return file_updates


def add_float_to_valuedata(extracted_adf: Adf, value: float) -> list[dict]:
  logger.debug(f"  Adding value {value} to ValueData array")
  valuedata = extracted_adf.table_instance_full_values[0].value["ValueData"]
  value_bytes = mods.create_bytearray(value, "float32")
  file_updates = []
  # update headers and instance offsets
  file_updates.extend(mods.update_non_instance_offsets(extracted_adf, len(value_bytes)))
  file_updates.extend(update_instance_offsets(extracted_adf, len(value_bytes), "ValueData"))
  # add to array and increase array length in header
  new_value_offset = valuedata.data_offset + (len(valuedata.value) * len(value_bytes))
  file_updates.append({
    "offset": new_value_offset,
    "value": value_bytes,
    "transform": "insert",
  })
  valuedata.value.append(value)
  file_updates.append({
    "offset": valuedata.info_offset + 8,
    "value": 1,
    "transform": "add"
  })
  return file_updates


def copy_string(old_string: AdfValue, str_bytes: bytearray) -> AdfValue:
  new_string = copy.deepcopy(old_string)
  new_string.value = str_bytes
  # find the end offset of the old string, add 1 for padding, and get the next offset divisible by 8 to ensure byte alignment
  old_string_end_offset = old_string.data_offset + len(old_string.value)
  new_string_offset = math.ceil((old_string_end_offset + 1) / 8) * 8
  new_string.info_offset += 8  # info header uint32 + 4-byte padding
  new_string.data_offset = new_string_offset + 8
  return new_string


def add_string_to_stringdata(extracted_adf: Adf, value: str) -> list[dict]:
  stringdata = extracted_adf.table_instance_full_values[0].value["StringData"]
  logger.debug(f"  Adding value {value} to StringData Array at index {len(stringdata.value)}")
  str_bytearray = mods.create_bytearray(value, "string")
  # Strings are stored as AdfValue objects. Copy the last string and update the values and offsets
  new_string = copy_string(stringdata.value[-1], value)
  file_updates = []
  # update headers and instance offsets
  added_size = len(str_bytearray) + 8  # padded byte length + info offset
  file_updates.extend(mods.update_non_instance_offsets(extracted_adf, added_size))
  file_updates.extend(update_instance_offsets(extracted_adf, added_size, "StringData"))
  # increase all string data offsets by 8 bytes for the new info offset in stringdata headers
  file_updates.extend(update_stringdata_offsets(extracted_adf, 8, index=len(stringdata.value)+1))
  logger.debug(f"  Writing string offset {new_string.data_offset} at offset {new_string.info_offset}")
  logger.debug(f"  Writing string value {bytes(str_bytearray)} at offset {new_string.data_offset}")
  # info offsets are separated by a uint32 of "8" (08 00 00 00)
  # there are also 4 null bytes before the first string
  # add uint32 separator before the new info offset and insert it before the null padding
  # string data_offsets are aligned to BOF in Adf object but aligned to instance_offset in bin
  instance_offset = extracted_adf.table_instance[0].offset
  file_updates.append({
    "offset": new_string.info_offset - 4,
    "value": mods.create_bytearray([8, new_string.data_offset - instance_offset], "uint32"),
    "transform": "insert"
  })
  # add to array and increase array length in header
  file_updates.append({
    "offset": new_string.data_offset,
    "value": str_bytearray,
    "transform": "insert",
  })
  stringdata.value.append(new_string)
  file_updates.append({
    "offset": stringdata.info_offset + 8,
    "value": 1,
    "transform": "add"
  })
  return file_updates


def copy_cell_definition(old_cell_def: AdfValue, cell: XlsxCell, value_index: int) -> AdfValue:
  cell_def = copy.deepcopy(old_cell_def)
  # Overwrite values with our desired data. A cell definition is 12 bytes long (uint16, 2 bytes padding, uint32, uint32)
  cell_def.data_offset += 12
  cell_def.info_offset += 12
  cell_def.value["Type"].value = cell.desired_data_type
  cell_def.value["Type"].data_offset += 12
  cell_def.value["Type"].info_offset += 12
  cell_def.value["DataIndex"].value = value_index
  cell_def.value["DataIndex"].data_offset += 12
  cell_def.value["DataIndex"].info_offset += 12
  cell_def.value["AttributeIndex"].value = cell.attribute_index
  cell_def.value["AttributeIndex"].data_offset += 12
  cell_def.value["AttributeIndex"].info_offset += 12
  return cell_def


def add_cell_definition(extracted_adf: Adf, cell: XlsxCell, value_index: int) -> list[dict]:
  logger.debug(f"  Adding new cell definition (Type: {cell.desired_data_type}  DataIndex: {value_index}  AttributeIndex: {cell.attribute_index})")
  cell_def_array = extracted_adf.table_instance_full_values[0].value["Cell"]
  # Cell definitions are AdfValue objects. Copy the last cell definition and update the values and offsets
  cell_def = copy_cell_definition(cell_def_array.value[-1], cell, value_index)
  cell_def_bytes = mods.create_bytearray([cell_def], "cell_definition")
  added_size = len(cell_def_bytes)
  file_updates = []
  # update headers and instance offsets
  file_updates.extend(mods.update_non_instance_offsets(extracted_adf, added_size))
  file_updates.extend(update_instance_offsets(extracted_adf, added_size, "Cell"))
  # add to array and increase array length in header
  file_updates.append({
    "offset": cell_def.data_offset,
    "value": cell_def_bytes,
    "transform": "insert",
  })
  cell_def_array.value.append(cell_def)
  file_updates.append({
    "offset": cell_def_array.info_offset + 8,
    "value": 1,
    "transform": "add"
  })
  return file_updates


def apply_coordinate_updates_to_file(src_filename: str, coordinate_updates: list[dict], skip_add_data: bool = False, allow_new_data: bool = False, force: bool = False) -> None:
  extracted_adf = deserialize_adf(src_filename)
  file_updates = []
  for coordinate_update in coordinate_updates:
    cell = XlsxCell(src_filename, extracted_adf, coordinate_update)
    allow_new_data = coordinate_update.get("allow_new_data", allow_new_data)
    file_updates.extend(process_cell_update(cell, extracted_adf, skip_add_data=skip_add_data, allow_new_data=allow_new_data, force=force))
  mods.apply_updates_to_file(src_filename, file_updates)


def update_file_at_coordinates(src_filename: str, coordinate_update: dict, skip_add_data: bool = False, allow_new_data: bool = False, force: bool = False) -> None:
  extracted_adf = deserialize_adf(src_filename)
  cell = XlsxCell(src_filename, extracted_adf, coordinate_update)
  file_updates = process_cell_update(cell, extracted_adf, skip_add_data=skip_add_data, allow_new_data=allow_new_data, force=force)
  mods.apply_updates_to_file(src_filename, file_updates)


def update_file_at_multiple_coordinates_with_value(src_filename: str, sheet_name: str, coordinates_list: list[str], value: any, transform: str = None, skip_add_data: bool = False, allow_new_data: bool = False, force: bool = False) -> None:
  coordinate_updates = [{
    "sheet": sheet_name,
    "coordinates": coordinates,
    "value": value,
    "transform": transform,
    "allow_new_data": allow_new_data,
    "skip_add_data": skip_add_data,
    "force": force
  } for coordinates in coordinates_list]
  apply_coordinate_updates_to_file(src_filename, coordinate_updates, skip_add_data=skip_add_data, allow_new_data=allow_new_data, force=force)


def get_data_array_for_data_type(extracted_adf: AdfValue, data_type: int) -> tuple[list, int]:
  if data_type == 0:
    array_name = "BoolData"
  if data_type == 1:
    array_name = "StringData"
  if data_type == 2:
    array_name = "ValueData"
  return array_name, extracted_adf[array_name].value, extracted_adf[array_name].data_offset


def find_cells(
    adf_values: dict[str, AdfValue],
    definition_indexes: list[int] = None,
    data_type: int = None,
    value_index: int = None,
    ignore_cell: XlsxCell = None,
  ) -> list[tuple[int, int, int]]:  # (sheet_index, cell_index, definition_index)
  matching_cells = []
  if definition_indexes:  # find cells with same definition index
    for sheet_index, sheet in enumerate(adf_values["Sheet"].value):
      matching_cells.extend([(sheet_index, cell_index, def_index) for cell_index, def_index in enumerate(sheet.value["CellIndex"].value) if def_index in definition_indexes])
  if data_type is not None and value_index is not None:  # find cells that point at the same value index
    defs_with_value_match = find_cell_definitions(adf_values, data_type, [value_index])
    matching_cells.extend(find_cells(adf_values, definition_indexes=[i for (i,_d) in defs_with_value_match]))
  if ignore_cell:  # don't return the cell we're trying to match
    matching_cells = [c for c in matching_cells if c != (ignore_cell.sheet_index, ignore_cell.index, ignore_cell.definition_index)]
  return matching_cells


def find_cell_definitions(
    adf_values: dict[str, AdfValue],
    data_type: int,
    value_indexes: list[int],
  ) -> list[tuple[int, dict]]:  # (definition_index, definition_object.value)
  cell_definitions = adf_values["Cell"].value
  defs_with_value_match = [
    (i, d.value)
    for i,d in enumerate(cell_definitions)
    if d.value["Type"].value == data_type and d.value["DataIndex"].value in value_indexes
  ]
  logger.debug(f"  Found {len(defs_with_value_match)} definitions with data type {data_type} pointing at value indexes in {value_indexes}")
  return defs_with_value_match


def get_unused_cell_def_indexes(extracted_adf: AdfValue) -> list[int]:  # defitition_index
  unused_definition_indexes = set(range(len(extracted_adf["Cell"].value)))
  for sheet in extracted_adf["Sheet"].value:
    in_use = set(sheet.value["CellIndex"].value)
    unused_definition_indexes.difference_update(in_use)
  return list(unused_definition_indexes)


def get_unused_values(adf_values: dict[str, AdfValue], data_array_name: str, data_type: int) -> list[dict]:
  all_value_indexes = list(range(len(adf_values[data_array_name].value)))
  unused_value_indexes = []
  for value_index in all_value_indexes:
     # check if any cells point at the index
    if not find_cells(adf_values, data_type=data_type, value_index=value_index):
      unused_value_indexes.append(value_index)

  unused_values = []
  for index in unused_value_indexes:
    if data_array_name == "StringData":
      value = adf_values[data_array_name].value[index].value
      offset = adf_values[data_array_name].value[index].data_offset
    else:
      value = adf_values[data_array_name].value[index]
      offset = adf_values[data_array_name].data_offset + (index * 4)
    unused_values.append({"index": index, "value": value, "offset": offset})
  return unused_values


def find_closest_value(value_array: list[float], desired_value: float) -> tuple[int, float]:
  match_index = None
  closest_delta = 9999999
  closest_match_index = None
  for i, number in enumerate(value_array):
    if float(number) == desired_value:
      match_index = i
      break
    delta = abs(float(number) - desired_value)
    if delta < closest_delta:
      closest_match_index = i
      closest_delta = delta
  if match_index is not None:
    return match_index, value_array[match_index]
  return closest_match_index, value_array[closest_match_index]


def calculate_coordinates(index: int, cols: int) -> str:
  cell_number = index + 1  # index starts at 0, cells start at A1
  row = math.ceil(cell_number / cols)
  col = cell_number - ( ( row - 1 ) * cols )
  return f"{get_column_letter(col)}{row}"


def range_to_coordinates_list(column: str, start_row: int, end_row: int) -> list[str]:
  coordinates_list = [f"{column}{row}" for row in range(start_row, end_row + 1)]
  return coordinates_list


def get_column_range(col_start: str, col_end: str) -> list[str]:
  start_index = column_index_from_string(col_start)
  end_index = column_index_from_string(col_end)
  return [get_column_letter(i) for i in range(start_index, end_index + 1)]


def get_coordinates_range_from_file(src_filename: str, sheet_name: str, rows: tuple[int, int] = (None, None), cols: tuple[str, str] = {None, None}) -> list[str]:
  extracted_adf = deserialize_adf(src_filename)
  sheet, _i = get_sheet(extracted_adf, sheet_name)
  if sheet is None:
    raise ValueError(f'Unable to find sheet "{sheet_name}" in file "{src_filename}"')
  row_start = rows[0] or 1
  row_end = rows[1] or sheet["Rows"].value  # default to last row in sheet
  row_range = range(row_start, row_end + 1)
  col_start = cols[0] or "A"
  col_end = cols[1] or get_column_letter(sheet["Cols"].value)  # default to last column in sheet
  column_range = get_column_range(col_start, col_end)
  return [f"{col}{row}" for col in column_range for row in row_range]

def least_sigfig(value):
  if value == int(value):
    return int(value)
  return float(f"{value:10g}")
